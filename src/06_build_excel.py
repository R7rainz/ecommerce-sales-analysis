"""Step 4c - The Excel analysis workbook.

Produces a multi-sheet .xlsx with the KPI summary, every insight table, and
native Excel charts (real chart objects, not pasted images, so they stay live
if the underlying figures are edited).

Run:  python src/06_build_excel.py
"""

import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import CLEAN_CSV, OUTPUT_DIR, PROCESSED_DIR, TABLE_DIR, ensure_dirs

WORKBOOK = OUTPUT_DIR / "ecommerce_analysis.xlsx"

# Matches the validated chart palette used elsewhere in the project.
BLUE = "2A78D6"
INK = "0B0B0B"
INK_SOFT = "52514E"
SURFACE_ALT = "F2F5FA"
RULE = "E6E5E1"

TITLE_FONT = Font(size=15, bold=True, color=INK)
SUB_FONT = Font(size=10, color=INK_SOFT)
HEAD_FONT = Font(size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor=BLUE)
THIN = Side(style="thin", color=RULE)
BORDER = Border(bottom=THIN)

MONEY = '#,##0.00'
PCT = '0.0%'
INT = '#,##0'


def sheet_title(ws, title: str, subtitle: str = "") -> int:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUB_FONT
        ws["A2"].alignment = Alignment(wrap_text=False)
        row = 3
    return row + 1


def write_table(ws, df: pd.DataFrame, start_row: int, formats: dict | None = None) -> int:
    """Write a dataframe as a formatted table. Returns the first row after it."""
    formats = formats or {}

    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col.replace("_", " "))
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, (_, record) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=i, column=j, value=record[col])
            cell.border = BORDER
            if col in formats:
                cell.number_format = formats[col]
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=SURFACE_ALT)

    for j, col in enumerate(df.columns, start=1):
        width = max(len(str(col)) + 2,
                    *(len(f"{v}") + 2 for v in df[col].head(60)))
        ws.column_dimensions[get_column_letter(j)].width = min(max(width, 11), 34)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(df) + 2


def style_chart(chart, title: str, height=8.5, width=17) -> None:
    chart.title = title
    chart.height = height
    chart.width = width
    chart.style = 2


def add_bar_chart(ws, df, anchor, title, cat_col, val_col, start_row):
    """Single-series bar - one colour for every bar (nominal categories)."""
    chart = BarChart()
    chart.type = "col"
    chart.gapWidth = 60
    style_chart(chart, title)

    ci = list(df.columns).index(cat_col) + 1
    vi = list(df.columns).index(val_col) + 1
    data = Reference(ws, min_col=vi, min_row=start_row, max_row=start_row + len(df))
    cats = Reference(ws, min_col=ci, min_row=start_row + 1, max_row=start_row + len(df))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.title = "Sales (Rs)"
    chart.legend = None
    for series in chart.series:
        series.graphicalProperties.solidFill = BLUE
        series.graphicalProperties.line.solidFill = BLUE
    ws.add_chart(chart, anchor)


def build_kpi_sheet(wb, kpis: dict, answers: dict, sig: pd.DataFrame) -> None:
    ws = wb.create_sheet("KPI Dashboard")
    ws.sheet_view.showGridLines = False
    row = sheet_title(
        ws,
        "E-Commerce Sales - Executive Summary",
        f"{kpis['total_orders']:,} orders | {kpis['date_from']} to {kpis['date_to']}",
    )

    cards = [
        ("Total Sales", kpis["total_sales"], MONEY),
        ("Total Profit", kpis["total_profit"], MONEY),
        ("Total Orders", kpis["total_orders"], INT),
        ("Total Quantity Sold", kpis["total_quantity"], INT),
        ("Average Discount", kpis["average_discount"], PCT),
        ("Profit Margin", kpis["overall_profit_margin"], PCT),
        ("Unique Customers", kpis["unique_customers"], INT),
        ("Avg Order Value", kpis["average_order_value"], MONEY),
    ]

    # KPI cards laid out two per row: label above, value below.
    for i, (label, value, fmt) in enumerate(cards):
        r = row + (i // 2) * 3
        c = 1 + (i % 2) * 3
        lab = ws.cell(row=r, column=c, value=label)
        lab.font = Font(size=9.5, bold=True, color=INK_SOFT)
        val = ws.cell(row=r + 1, column=c, value=value)
        val.font = Font(size=18, bold=True, color=BLUE)
        val.number_format = fmt

    row = row + ((len(cards) + 1) // 2) * 3 + 1

    ws.cell(row=row, column=1, value="Answers to the five business questions").font = \
        Font(size=12, bold=True, color=INK)
    row += 1
    ws.cell(row=row, column=1,
            value="Each answer is the literal top of its ranking. The Confidence "
                  "column reports whether that ranking survives statistical testing.").font = SUB_FONT
    row += 2

    retention = dict(zip(sig["Dimension"], sig["Leader_Retention_Rate"]))
    qa = pd.DataFrame([
        {"Question": "1. Highest-sales state",
         "Answer": answers["q1_highest_sales_state"]["answer"],
         "Figure": f"Rs {answers['q1_highest_sales_state']['total_sales']:,.0f}",
         "Confidence": f"NOT SIGNIFICANT - holds top in only "
                       f"{retention.get('State', 0):.0%} of resamples"},
        {"Question": "2. Most profitable category",
         "Answer": answers["q2_most_profitable_category"]["answer"],
         "Figure": f"Rs {answers['q2_most_profitable_category']['total_profit']:,.0f}",
         "Confidence": f"NOT SIGNIFICANT - holds top in "
                       f"{retention.get('Product_Category', 0):.0%} of resamples"},
        {"Question": "3. Highest-revenue product",
         "Answer": answers["q3_highest_revenue_products"]["answer"],
         "Figure": f"Rs {answers['q3_highest_revenue_products']['total_sales']:,.0f}",
         "Confidence": f"NOT SIGNIFICANT - holds top in "
                       f"{retention.get('Product_Name', 0):.0%} of resamples"},
        {"Question": "4. Most-used payment mode",
         "Answer": answers["q4_most_used_payment_method"]["answer"],
         "Figure": f"{answers['q4_most_used_payment_method']['share_of_orders']:.1%} of orders",
         "Confidence": f"NOT SIGNIFICANT - even split is ~20.0% per mode"},
        {"Question": "5. Monthly sales trend",
         "Answer": "Flat - no trend",
         "Figure": f"H2 vs H1 {answers['q5_monthly_sales_trend']['first_half_vs_second_half_change']:+.1%}",
         "Confidence": "NO TREND - linear fit R2 = 0.008"},
    ])
    write_table(ws, qa, row)


def main() -> None:
    ensure_dirs()
    print("=" * 70)
    print("STEP 4c - EXCEL WORKBOOK")
    print("=" * 70 + "\n")

    payload = json.loads((OUTPUT_DIR / "analysis_summary.json").read_text())
    kpis, answers = payload["kpis"], payload["answers"]
    sig = pd.read_csv(TABLE_DIR / "significance_tests.csv")

    wb = Workbook()
    wb.remove(wb.active)

    build_kpi_sheet(wb, kpis, answers, sig)
    print("  built KPI Dashboard")

    money_fmt = {
        "Total_Sales": MONEY, "Total_Profit": MONEY, "Avg_Discount": PCT,
        "Profit_Margin": PCT, "Sales_Share": PCT, "Total_Orders": INT,
        "Total_Quantity": INT, "MoM_Growth": PCT,
    }

    specs = [
        ("sales_by_state", "Sales by State", "State",
         "Every state falls within 4% of the mean - the ranking is not meaningful."),
        ("sales_by_category", "Sales by Category", "Product_Category",
         "All five categories sit within 5% of each other."),
        ("profit_by_category", "Profit by Category", "Product_Category",
         "Margins are ~15% everywhere; no category is meaningfully more profitable."),
        ("top_10_products", "Top 10 Products", "Product_Name",
         "Spread across the top ten is under 12%, consistent with uniform demand."),
        ("sales_by_payment_mode", "Payment Modes", "Payment_Mode",
         "Five modes at roughly 20% of orders each - an even split."),
        ("sales_by_city", "Sales by City", "City",
         "City is independent of State in this dataset - never drill down from one to the other."),
    ]

    for name, sheet, cat_col, note in specs:
        df = pd.read_csv(TABLE_DIR / f"{name}.csv")
        ws = wb.create_sheet(sheet)
        ws.sheet_view.showGridLines = False
        start = sheet_title(ws, sheet, note)
        after = write_table(ws, df, start, money_fmt)
        add_bar_chart(ws, df, f"A{after + 1}", f"{sheet} (Rs)",
                      cat_col, "Total_Sales", start)
        print(f"  built {sheet}")

    # Monthly trend gets a line chart with a zero-based axis.
    trend = pd.read_csv(TABLE_DIR / "monthly_sales_trend.csv")
    ws = wb.create_sheet("Monthly Trend")
    ws.sheet_view.showGridLines = False
    start = sheet_title(
        ws, "Monthly Sales Trend",
        "Zero-based axis. Sales are flat across 24 months (linear fit R2 = 0.008).",
    )
    after = write_table(ws, trend, start, money_fmt)

    chart = LineChart()
    style_chart(chart, "Monthly Sales (Rs)")
    data = Reference(ws, min_col=2, min_row=start, max_row=start + len(trend))
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(trend))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.scaling.min = 0  # never truncate a flat series
    chart.y_axis.title = "Sales (Rs)"
    chart.legend = None
    chart.series[0].graphicalProperties.line.solidFill = BLUE
    chart.series[0].graphicalProperties.line.width = 22000
    chart.series[0].smooth = False
    ws.add_chart(chart, f"A{after + 1}")
    print("  built Monthly Trend")

    # Significance evidence.
    ws = wb.create_sheet("Significance Tests")
    ws.sheet_view.showGridLines = False
    start = sheet_title(
        ws, "Are the rankings real?",
        "Permutation tests and bootstrap confidence intervals. Every leader's "
        "margin CI crosses zero, so no ranking is distinguishable from noise.",
    )
    write_table(ws, sig, start, {
        "Observed_Margin": MONEY, "Margin_CI_Low": MONEY,
        "Margin_CI_High": MONEY, "Leader_Retention_Rate": PCT,
        "Permutation_P_Value": '0.0000',
    })
    print("  built Significance Tests")

    # Full cleaned data last - big sheet, so it does not slow down navigation.
    clean = pd.read_csv(CLEAN_CSV, parse_dates=["Order_Date"])
    ws = wb.create_sheet("Cleaned Data")
    ws.append(list(clean.columns))
    for cell in ws[1]:
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
    for record in clean.itertuples(index=False):
        ws.append(list(record))
    for j, col in enumerate(clean.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(len(col) + 3, 13)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    print(f"  built Cleaned Data ({len(clean):,} rows)")

    wb.save(WORKBOOK)
    print(f"\nWrote {WORKBOOK.relative_to(PROCESSED_DIR.parent.parent)}")
    print(f"  {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
